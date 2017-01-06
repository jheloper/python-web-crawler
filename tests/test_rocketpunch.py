import unittest
import openpyxl
from datetime import datetime

from src.spider import RocketpuchSpider


class RocketpunchTest(unittest.TestCase):
    def setUp(self):
        self.spider = RocketpuchSpider(job='개발자', specialty='python')

    def test_connect(self):
        self.assertEqual(200, self.spider.response.status_code)
        self.assertIn('로켓펀치', self.spider.beautiful_soup.title.text)

    def test_connect_with_parameter(self):
        self.assertEqual(200, self.spider.response.status_code)
        self.assertIn('개발자', self.spider.beautiful_soup.title.text)
        self.assertIn('Python', self.spider.beautiful_soup.title.text)

    def test_get_pages(self):
        self.assertTrue(self.spider.get_max_page() > 0)

    def test_get_job_info(self):
        job_item_list = self.spider.get_job_info()
        self.assertEqual(0, len([job_item for job_item in job_item_list if not job_item.job_title]))

    def test_make_excel(self):
        wb = openpyxl.Workbook()
        ws_title = datetime.today().strftime('%Y.%m.%d')
        ws = wb.create_sheet(index=0, title=ws_title)
        ws['A1'] = 'test'
        wb.save(filename='jobs.xlsx')
        wb2 = openpyxl.load_workbook('jobs.xlsx')
        ws2 = wb2[ws_title]
        self.assertEqual('test', ws2['A1'].value)

    def test_convert_to_excel(self):
        self.fail()


if __name__ == '__main__':
    unittest.main()
